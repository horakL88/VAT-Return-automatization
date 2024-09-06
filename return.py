from openpyxl import load_workbook
import warnings
import pandas as pd
from lxml import etree as et

warnings.simplefilter(action='ignore', category=UserWarning)

file_path = r'VAT_RETURN_INPUT_template.xlsx'
# Excel worksheetet aktiválom
wb = load_workbook(file_path)
ws = wb.active

info = wb['INFO']
og_output = wb['VAT output']
og_input = wb['VAT input']

# OUTPUT oldal
output = wb.create_sheet("OUTPUT")

OPdf = pd.DataFrame(wb['VAT output'].values)
IPdf = pd.DataFrame(wb['VAT input'].values)
#OPdf.reset_index(drop=True) # Resets the index, makes factor a column
#OPdf.drop("Factor",axis=1,inplace=True) # drop factor from axis 1 and make changes permanent by inplace=True
#print(OPdf.head())
#print (OPdf.columns.tolist())
#OPdf.loc[OPdf['a'] == 1, 'b'].sum()



# XML készítése

root = et.Element('xml', version = "1.0", encoding="UTF-8")

nyomtatvanyok = et.Element('nyomtatvanyok', xmlns = "http://www.apeh.hu/abev/nyomtatvanyok/2005/01")


nyomtatvany = et.SubElement(nyomtatvanyok, 'nyomtatvany')
nyomtatvanyinformacio = et.SubElement(nyomtatvany, 'nyomtatvanyinformacio')
nyomtatvanyazonosito = et.SubElement(nyomtatvanyinformacio, 'nyomtatvanyazonosito')
nyomtatvanyazonosito.text = "2465A"
nyomtatvanyverzio = et.SubElement(nyomtatvanyinformacio, 'nyomtatvanyverzio')
nyomtatvanyverzio.text = "1.0"

# Adózó azonosítói
adozo = et.SubElement(nyomtatvanyinformacio, 'adozo')

nev = et.SubElement(adozo, 'nev')
if info['B3'].value is None:
    print("FIGYELEM: Adózó neve hiányzik!")
else:
    nev.text = info['B3'].value

adoszam = et.SubElement(adozo, 'adoszam')
if info['B4'].value is None:
    print("FIGYELEM: Adózó adószáma hiányzik!")
else:
    adoszam.text = str(info['B4'].value)

# Időszak meghatározása
idoszak = et.SubElement(nyomtatvanyinformacio, 'idoszak')

tol = et.SubElement(idoszak, 'tol')
if info['B8'].value is None:
    pass
else:
    tol.text = str(info['B8'].value)

ig = et.SubElement(idoszak, 'ig')
if info['B9'].value is None:
    pass
else:
    ig.text = str(info['B9'].value)

# Információs Mezők
mezok = et.SubElement(nyomtatvany, 'mezok')

adoszam_mezo = et.SubElement(mezok, 'mezo', eazon = "0A0001E001A")
if info['B4'].value is None:
    print("FIGYELEM: Adózó adószáma hiányzik!")
else:
    adoszam_mezo.text = str(info['B4'].value)

nev_mezo = et.SubElement(mezok, 'mezo', eazon = "0A0001E006A")
if info['B3'].value is None:
    print("FIGYELEM: Adózó neve hiányzik!")
else:
    nev_mezo.text = info['B3'].value

ugyintezo_neve = et.SubElement(mezok, 'mezo', eazon = "0A0001E007A")
if info['B5'].value is None:
    pass
else:
    ugyintezo_neve.text = info['B5'].value

ugyintezo_telefonszama = et.SubElement(mezok, 'mezo', eazon = "0A0001E008A")
if info['B6'].value is None:
    pass
else:
    ugyintezo_telefonszama.text = info['B6'].value

tol_mezo = et.SubElement(mezok, 'mezo', eazon = "0A0001F001A")
if info['B8'].value is None:
    pass
else:
    tol_mezo.text = str(info['B8'].value)

ig_mezo = et.SubElement(mezok, 'mezo', eazon = "0A0001F002A")
if info['B9'].value is None:
    pass
else:
    ig_mezo.text = str(info['B9'].value)

# Bevallas jellege
bevallas_jellege = et.SubElement(mezok, 'mezo', eazon = "0A0001F005A")
if info['B10'].value is None:
    pass
else:
    bevallas_jellege.text = str(info['B10'].value)

bevall_gyakori = et.SubElement(mezok, 'mezo', eazon = "0A0001F006A")
if info['B7'].value is None:
    pass
else:
    bevall_gyakori.text = str(info['B7'].value)


# Adózó belföldi pénzforgalmi vagy fizetési számlaszáma 1
belf_penzforg_szamlaszam1 = et.SubElement(mezok, 'mezo', eazon = "0A0001G001A")
if info['B18'].value is None:
    pass
else:
    belf_penzforg_szamlaszam1.text = str(info['B18'].value)[:8]
# Adózó belföldi pénzforgalmi vagy fizetési számlaszáma 2
belf_penzforg_szamlaszam2 = et.SubElement(mezok, 'mezo', eazon = "0A0001G002A")
if info['B18'].value is None:
    pass
else:
    belf_penzforg_szamlaszam2.text = str(info['B18'].value)[9:16]
# Számlavezető pénzforgalmi szolgáltató neve
szamlavezeto_penzforg_nev = et.SubElement(mezok, 'mezo', eazon = "0A0001G003A")
#szamlavezeto_penzforg_nev.text = str("")
# Külföldi számla tulajdonosának neve
kulf_szamla_tulaj_nev = et.SubElement(mezok, 'mezo', eazon = "0A0001G007A")
if info['B18'].value is None:
    pass
else:
    kulf_szamla_tulaj_nev.text = str(info['B27'].value)
# IBAN szám jelölése
IBAN_no = et.SubElement(mezok, 'mezo', eazon = "0A0001G008A")
if info['B23'].value == None:
    pass
else:
    IBAN_no.text = str(info['B23'].value)
# SWIFT kód
SWIFT_code = et.SubElement(mezok, 'mezo', eazon = "0A0001G009A")
if info['B22'].value == None:
    pass
else:
    SWIFT_code.text = str(info['B22'].value)
# Adózó külföldi fizetési számlaszáma
adozo_kulf_szamlaszam = et.SubElement(mezok, 'mezo', eazon = "0A0001G010A")
if info['B21'].value == None:
    pass
else:
    adozo_kulf_szamlaszam.text = str(info['B21'].value)
# Országkód
orszagkod = et.SubElement(mezok, 'mezo', eazon = "0A0001G011A")
if info['B26'].value == None:
    pass
else:
    orszagkod.text = str(info['B26'].value)
# Devizanem
devizanem = et.SubElement(mezok, 'mezo', eazon = "0A0001G012A")
if info['B28'].value == None:
    pass
else:
    devizanem.text = str(info['B28'].value)
# Külföldi számlát vezető pénzügyi szolgáltató neve
kulf_penz_szolg_nev = et.SubElement(mezok, 'mezo', eazon = "0A0001G013A")
#kulf_penz_szolg_nev.text = str("")
# Külföldi számlát vezető pénzügyi szolgáltató címe
kulf_penz_szolg_cim = et.SubElement(mezok, 'mezo', eazon = "0A0001G014A")
#kulf_penz_szolg_cim.text = str("")
# Teljes összeg kiutalása kérésének jelölése
teljes_osszeg_kiutalas_keres = et.SubElement(mezok, 'mezo', eazon = "0A0001G015A")
if info['B29'].value == None:
    pass
else:
    teljes_osszeg_kiutalas_keres.text = str(info['B29'].value)
# Kiutalást nem kér jelölése
kiutalas_nem_ker = et.SubElement(mezok, 'mezo', eazon = "0A0001G016A")
if info['B31'].value == None:
    pass
else:
    kiutalas_nem_ker.text = str(info['B31'].value)
# Átvezetési és kiutalási kérelem mellékelve
atvezetes_ker = et.SubElement(mezok, 'mezo', eazon = "0A0001G017A")
if info['B32'].value == None:
    pass
else:
    atvezetes_ker.text = str(info['B32'].value)
# Adózó nyilvánosan működő részvénytársaság
nyrt = et.SubElement(mezok, 'mezo', eazon = "0A0001G018A")
#nyrt.text = str("")
# Nyilatkozat az Art. 64. § (3) bekezdése szerint
Art64 = et.SubElement(mezok, 'mezo', eazon = "0A0001G020A")
if info['B33'].value == None:
    pass
else:
    Art64.text = str(info['B33'].value)
# Felelősségem tudatában kijelentem, hogy a bevallásban közölt adatok a valóságnak megfelelnek
helyseg = et.SubElement(mezok, 'mezo', eazon = "0A0001I001A")
if info['B14'].value == None:
    pass
else:
    helyseg.text = info['B14'].value
# Bevallás beadásának dátuma
bevall_datum = et.SubElement(mezok, 'mezo', eazon = "0A0001I002A")
if info['B15'].value == None:
    pass
else:
    bevall_datum.text = str(info['B15'].value)
# állandó meghatalmazás jelölése
allando_meghat = et.SubElement(mezok, 'mezo', eazon = "0A0001I005A")
if info['B11'].value == None:
    pass
else:
    allando_meghat.text = info['B11'].value

# Eseti meghatalmazás csatolásának jelölése
eseti_meghat = et.SubElement(mezok, 'mezo', eazon = "0A0001I006A")
if info['B13'] is None:
    pass
else:
    eseti_meghat.text = info['B13'].value



# 2465A 01-01

# Közösség területén kívülre történő termékértékesítés, azzal egy tekintet alá eső szolgáltatásnyújtás, 
# valamint nemzetközi közlekedéshez kapcsolódó termékértékesítés és szolgáltatásnyújtás - 0B0001C0001BA H9 Export
export_value = OPdf.loc[OPdf[2] == 'H9', 17].sum()

print ("Export: ")
print (int(round(export_value, -3)/1000))

export = et.SubElement(mezok, 'mezo',eazon = "0B0001C0001BA")
export.text = str(int(round(export_value, -3)/1000))

# Közösségen belülre történő, adólevonási joggal járó adómentes termékértékesítés (kivéve az új közlekedési eszköz értékesítését)
# Intra Community sales
# 0B0001C0002BA
intra_com_sales_value = OPdf.loc[OPdf[2] == 'HE', 17].sum()

print ("Intra Community sales: ")
print (int(round(intra_com_sales_value, -3)/1000))

intra_com_sales = et.SubElement(mezok, 'mezo', eazon = "0B0001C0002BA")
intra_com_sales.text = str(int(round(export_value, -3)/1000))

# Új közlekedési eszköz Közösségen belülre történő értékesítésének összege
# 0B0001C0003BA

# Áfa tv. 142 §-a szerint termékértékesítés, szolgáltatásnyújtás és az adólevonással járó adómentes belföldi értékesítés ellenértéke
# 0B0001C0004BA

# 0 %-os kulcs alá tartozó értékesítés
# 0B0001C0110BA

# 5 %-os kulcs alá tartozó értékesítés
# Domestic Sale - 5% NET - 0B0001C0005BA VAT - 0B0001C0005CA
domsale_5_VAT_value = OPdf.loc[OPdf[2] == 'H6', 17].sum()
if domsale_5_VAT_value < 0:
    domsale_5_VAT_value = domsale_5_VAT_value * -1
domsale_5_NET_value = domsale_5_VAT_value / 0.05


print ("Domestic Sales - 5%: ")
print ("NET: ", int(round(domsale_5_NET_value, -3)/1000), " VAT: ", int(round(domsale_5_VAT_value,-3)/1000))

domsale_5_NET = et.SubElement(mezok, 'mezo',eazon = "0B0001C0005BA")
domsale_5_NET.text = str(int(round(domsale_5_NET_value, -3)/1000))

domsale_5_VAT = et.SubElement(mezok, 'mezo',eazon = "0B0001C0005CA")
domsale_5_VAT.text = str(int(round(domsale_5_VAT_value, -3)/1000))

# 18 %-os kulcs alá tartozó értékesítés
# Domestic Sale - 18% NET - 0B0001C0006BA VAT - 0B0001C0006CA
domsale_18_VAT_value = OPdf.loc[OPdf[2] == 'UA', 17].sum()
if domsale_18_VAT_value < 0:
    domsale_18_VAT_value = domsale_18_VAT_value * -1
domsale_18_NET_value = domsale_18_VAT_value / 0.18

print ("Domestic Sales - 18%: ")
print ("NET: ", int(round(domsale_18_NET_value, -3)/1000), " VAT: ", int(round(domsale_18_VAT_value,-3)/1000))

domsale_18_NET = et.SubElement(mezok, 'mezo',eazon = "0B0001C0006BA")
domsale_18_NET.text = str(int(round(domsale_18_NET_value, -3)/1000))

domsale_18_VAT = et.SubElement(mezok, 'mezo',eazon = "0B0001C0006CA")
domsale_18_VAT.text = str(int(round(domsale_18_VAT_value, -3)/1000))

# 27 %-os kulcs alá tartozó értékesítés
# Domestic Sale - 27% NET - 0B0001C0007BA VAT - 0B0001C0007CA
domsale_27_VAT_value = OPdf.loc[OPdf[2] == 'HI', 17].sum()
if domsale_27_VAT_value < 0:
    domsale_27_VAT_value = domsale_27_VAT_value * -1
domsale_27_NET_value = domsale_27_VAT_value / 0.27

print ("Domestic Sales - 27%: ")
print ("NET: ", int(round(domsale_27_NET_value, -3)/1000), " VAT: ", int(round(domsale_27_VAT_value,-3)/1000))

domsale_27_NET = et.SubElement(mezok, 'mezo',eazon = "0B0001C0007BA")
domsale_27_NET.text = str(int(round(domsale_27_NET_value, -3)/1000))

domsale_27_VAT = et.SubElement(mezok, 'mezo',eazon = "0B0001C0007CA")
domsale_27_VAT.text = str(int(round(domsale_27_VAT_value, -3)/1000))


# Közérdekű vagy egyéb speciális jellegére tekintettel adómentes értékesítés
# 0B0001C0008BA

# Különleges eljárással megállapított adó
# NET: 0B0001C0009BA VAT: 0B0001C0009CA

# Saját vállalkozáson belül végzett beruházás után fizetendő adó
# NET: 0B0001C0010BA VAT: 0B0001C0010CA

# Közösségen belülről történő adómentes termékbeszerzés
# 0B0001C0011BA

# Közösségen belülről történő 0 %-os kulcs alá tartozó termékbeszerzés
# 0B0001C0112BA

# Közösségen belülről történő 5 %-os kulcs alá tartozó termékbeszerzés
# Intra-Community transfer - 5%
# NET: 0B0001C0012BA VAT: 0B0001C0012CA
intra_com_aq_VAT_value = IPdf.loc[IPdf[2] == 'WV', 20].sum()
intra_com_aq_NET_value = intra_com_aq_VAT_value / 0.05

print ("Intra Community Acquisitions - 5%: ")
print ("NET: ", int(round(intra_com_aq_NET_value, -3)/1000), "VAT: ", int(round(intra_com_aq_VAT_value,-3)/1000))

intra_com_aq_NET = et.SubElement(mezok, 'mezo', eazon = "0B0001C0012BA")
intra_com_aq_NET.text = str(int(round(intra_com_aq_NET_value,-3)/1000))

intra_com_aq_VAT = et.SubElement(mezok, 'mezo', eazon = "0B0001C0012CA")
intra_com_aq_VAT.text = str(int(round(intra_com_aq_VAT_value,-3)/1000))

# Közösségen belülről történő 18 %-os kulcs alá tartozó termékbeszerzés
# Intra-Community transfer - 18%
# NET: 0B0001C0013BA VAT: 0B0001C0013CA

intra_com_aq_VAT_value = IPdf.loc[IPdf[2] == 'WU', 20].sum()
intra_com_aq_NET_value = intra_com_aq_VAT_value / 0.18

print ("Intra Community Acquisitions - 18%: ")
print ("NET: ", int(round(intra_com_aq_NET_value, -3)/1000), "VAT: ", int(round(intra_com_aq_VAT_value,-3)/1000))

intra_com_aq_NET = et.SubElement(mezok, 'mezo', eazon = "0B0001C0013BA")
intra_com_aq_NET.text = str(int(round(intra_com_aq_NET_value,-3)/1000))

intra_com_aq_VAT = et.SubElement(mezok, 'mezo', eazon = "0B0001C0013CA")
intra_com_aq_VAT.text = str(int(round(intra_com_aq_VAT_value,-3)/1000))

# Közösségen belülről történő 27 %-os kulcs alá tartozó termékbeszerzés
# Intra-Community transfer - 27%
# NET: 0B0001C0014BA VAT: 0B0001C0014CA

intracom27_1 = IPdf.loc[IPdf[2] == 'WA', 20].sum()
intracom27_2 = IPdf.loc[IPdf[2] == 'UH', 20].sum()
intra_com_aq_VAT_value = intracom27_1 + intracom27_2
intra_com_aq_NET_value = intra_com_aq_VAT_value / 0.27

print ("Intra Community Acquisitions - 27%: ")
print ("NET: ", int(round(intra_com_aq_NET_value, -3)/1000), "VAT: ", int(round(intra_com_aq_VAT_value,-3)/1000))

intra_com_aq_NET = et.SubElement(mezok, 'mezo', eazon = "0B0001C0014BA")
intra_com_aq_NET.text = str(int(round(intra_com_aq_NET_value,-3)/1000))

intra_com_aq_VAT = et.SubElement(mezok, 'mezo', eazon = "0B0001C0014CA")
intra_com_aq_VAT.text = str(int(round(intra_com_aq_VAT_value,-3)/1000))

# Közösségen belülről történő új közlekedési eszköz beszerzés (27 %-os adómérték)
# NET: 0B0001C0015BA VAT: 0B0001C0015CA

# # Közösségen belülről történő jövedéki termékbeszerzés (27 %-os adómérték)
# NET: 0B0001C0016BA VAT: 0B0001C0016CA

# Adómentes szolgáltatás igénybevétel (közösségi adóalanytól és harmadik országbeli adóalanytól)
# 0B0001C0017BA

# Közösségi adóalanytól igénybe vett szolgáltatás utáni adófizetési kötelezettség az Áfa tv. 37. § (1) bekezdése alapján (27 %-os adómérték)
# NET: 0B0001C0018BA VAT: 0B0001C0018CA

# Közösségi adóalanytól igénybe vett szolgáltatás utáni egyéb adófizetési kötelezettség
# NET: 0B0001C0019BA VAT: 0B0001C0019CA

# Közösségen belül az Áfa tv. 91. § (2) bek. szerinti ügylet esetén a beszerző termékértékesítés címén  fizetendő 0 %-os mérték alá tartozó adóalap
# 0B0001C0113BA

# Közösségen belül az Áfa tv. 91. § (2) bek. szerinti ügylet esetén a beszerző termékértékesítés címén fizetendő 5%-os mértékű adója 
# NET: 0B0001C0020BA VAT: 0B0001C0020CA

# Közösségen belül az Áfa tv. 91. § (2) bek. szerinti ügylet esetén a beszerző termékértékesítés címén fizetendő 18 %-os mértékű adója
# NET: 0B0001C0021CA VAT: 0B0001C0021BA

# Közösségen belül az Áfa tv. 91. § (2) bek. szerinti ügylet esetén a beszerző termékértékesítés címén fizetendő 27 %-os mértékű adója
# NET: 0B0001C0022BA VAT: 0B0001C0022CA

# Adómentes termékimport
# 0B0001C0023BA

# 0 %-os mértékű termékimport
# 0B0001C0114BA

# Termékimport címén fizetendő 5 %-os mértékű adó
# NET: 0B0001C0024BA VAT: 0B0001C0024CA

# Termékimport címén fizetendő 18 %-os mértékű adó
# NET: 0B0001C0025BA VAT: 0B0001C0025CA

# Termékimport címén fizetendő 27 %-os mértékű adó
# NET: 0B0001C0026BA VAT: 0B0001C0026CA

# Harmadik országbeli adóalanytól igénybevett szolgáltatás utáni adófizetési kötelezettség
# NET: 0B0001C0027BA VAT: 0B0001C0027CA

# Az Áfa tv. 32. §, 34. § szerinti termékbeszerzés (27 %-os adómérték)
# NET: 0B0001C0028BA VAT: 0B0001C0028CA

# Az Áfa tv. 142. §-a alapján a fordított adózás szabályai szerint fizetendő adó
# NET: 0B0001C0029BA VAT: 0B0001C0029CA

# Az Áfa tv. 99. § (9) bekezdés alapján fizetendő adót csökkentő tétel
# NET: 0B0001C0030BA VAT: 0B0001C0030CA

# 31. Áfa tv. 153/C. § alapján fizetendőt növelő tétel összesen
# NET: 0B0001C0031BA VAT: 0B0001C0031CA

credit31_18 = IPdf.loc[(IPdf[2] == 'UC') & (IPdf[IPdf < 0]), 20].sum()

# df.loc[(df['a'] == 1) & (df['c'] == 2), 'b'].sum()

credit31_27 = IPdf.loc[(IPdf[2] == 'H2') & (IPdf[20][0] == "-" ), 20].sum()

credit31 = credit31_27 + credit31_18
print(credit31)


# 35. Egyéb
# NET: 0B0001C0035BA VAT: 0B0001C0035CA


OPdf.to_excel(r"C:\Users\EQ556AF\OneDrive - EY\Desktop\Programok\RETVRN\pandas_op.xlsx")
IPdf.to_excel(r"C:\Users\EQ556AF\OneDrive - EY\Desktop\Programok\RETVRN\pandas_ip.xlsx")

xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n' + et.tostring(nyomtatvanyok, pretty_print=True).decode("utf-8")
#print (xml_str)
with open (r'C:\Users\EQ556AF\OneDrive - EY\Desktop\Programok\RETVRN\test.xml', 'w',) as f:
    f.write(xml_str)